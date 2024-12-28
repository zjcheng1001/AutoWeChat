import os
import time
import datetime

import uiautomator2
from uiautomator2.exceptions import UiObjectNotFoundError
from uiautomator2.exceptions import RPCError

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment


# 计时装饰器
def time_decorator(func):
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        end_time = time.time()
        elapsed_time = end_time - start_time
        # 计算分钟和小时（如果需要）
        minutes, seconds = divmod(elapsed_time, 60)
        hours, minutes = divmod(minutes, 60)

        # 根据时间长度选择合适的时间格式进行打印
        if hours:
            print(f"{func.__name__}执行了{hours}小时{minutes:.0f}分钟{seconds:.2f}秒")
        elif minutes:
            print(f"{func.__name__}执行了{minutes:.0f}分钟{seconds:.2f}秒")
        else:
            print(f"{func.__name__}执行了{elapsed_time:.2f}秒")
        return result

    return wrapper


class WeChat:
    def __init__(self):
        self.driver = uiautomator2.connect()
        self.excel_path = f'微信好友_{datetime.datetime.now().strftime("%Y_%m_%d")}.xlsx'
        
    def start_wechat(self):
        self.driver.app_start("com.tencent.mm")

    @property
    def get_myself_name(self):
        self.driver(text='我').click()
        self.driver(className='android.widget.LinearLayout').child(className='android.view.View').wait(timeout=3)
        return self.driver(className='android.widget.LinearLayout').child(className='android.view.View').get_text()
    
    def get_current_view_all_person(self):
        """
        返回当前页面的所有联系人
        用xpath定位,用all()返回所有元素,这种方式速度更快(如果用resource-id定位,再一个一个用get_text()速度会慢很多)
        """
        return self.driver.xpath('//android.widget.TableRow/android.widget.TextView').all()
    
    def get_person_info(self, index, is_friend=True, save_to_excel=True):
        """
        获取人员信息
        :param index: 序号
        :param is_friend: 开关标志位，默认打开，查询是否为好友
        :param save_to_excel: 是否保存到Excel，默认为True
        :return: 一个包含个人信息的字典
        """
        # 确认新界面加载完成
        self.driver(text='发消息').wait(timeout=3)
        # 备注，昵称，微信号，地区，个性签名
        person_info_area = self.driver.xpath('//android.widget.TextView[starts-with(@text, "微信号")]/..')
        info_dict = {'序号': index, '备注': '', '昵称': '', '微信号': '', '地区': '', '个性签名': '',
                        '我和TA的共同群聊': '', '来源': '', '添加时间': '', '对方是否是好友': '',
                        '日期': datetime.datetime.now().strftime("%Y_%m_%d %H:%M:%S"),
                        '头像': ''}
        info_list = [person_info_area.child('//android.widget.LinearLayout/android.widget.TextView').all()[0].text] + [
            info.text for info in person_info_area.child('android.widget.TextView').all()]
        # 根据 info_list 的长度和内容填充剩余信息
        if len(info_list) == 4:
            info_dict['备注'], info_dict['昵称'], info_dict['微信号'], info_dict['地区'] = info_list
        elif len(info_list) == 3:
            if '微信号' in info_list[1]:
                info_dict['昵称'], info_dict['微信号'], info_dict['地区'] = info_list
            else:
                info_dict['备注'], info_dict['昵称'], info_dict['微信号'] = info_list
        elif len(info_list) == 2:
            info_dict['昵称'], info_dict['微信号'] = info_list
        # 格式化信息，去除多余的前缀
        info_dict['昵称'] = info_dict['昵称'].lstrip('昵称:  ') if '昵称' in info_dict['昵称'] else info_dict['昵称']
        info_dict['微信号'] = info_dict['微信号'].lstrip('微信号:  ') if '微信号' in info_dict['微信号'] else info_dict[
            '微信号']
        info_dict['地区'] = info_dict['地区'].lstrip('地区:  ') if '地区' in info_dict['地区'] else info_dict['地区']

        # 头像保存的路径
        head_dir = 'head_' + str(datetime.datetime.now().strftime("%Y_%m_%d"))
        if not os.path.exists(head_dir):
            os.mkdir(head_dir)
        head_name = info_dict['备注'] if info_dict['备注'] else info_dict['昵称']
        head_path = os.path.join(head_dir, head_name + '.png')
        # 对头像截图，保存在head_dir文件夹下
        max_retries = 3
        retries = 0
        while retries < max_retries:
            try:
                head = self.driver(description='头像')
                head.wait(timeout=3)
                head.screenshot().save(head_path)
                if os.path.exists(head_path):
                    info_dict['头像'] = head_path
                    break
            except RPCError:
                retries += 1
                time.sleep(1)  # 等待1秒后重试
                continue
            except Exception as e:
                print(f"保存头像失败: {e}")
                break

        if os.path.exists(head_path):
            info_dict['头像'] = head_path
            
        if self.driver(text='视频号').wait(timeout=0.5):    # 判断是否有视频号，如果有的话，微信会刷新页面，需要等待，不然点击更多信息时，会点击到视频号
            time.sleep(0.5)

        self.driver(text='更多信息').click()
        self.driver(text='来源').wait(timeout=3)
        signature = self.driver.xpath('//android.widget.TextView[starts-with(@text, "个性签名")]/../android.widget.TextView[2]').all()
        info_dict['个性签名'] = signature[0].text if signature else ''
        group_chat = self.driver.xpath('//android.widget.TextView[contains(@text, "共同群聊")]/../android.widget.TextView[2]').all()
        info_dict['我和TA的共同群聊'] = group_chat[0].text if group_chat else ''
        source = self.driver.xpath('//android.widget.TextView[starts-with(@text, "来源")]/../android.widget.TextView[2]').all()
        info_dict['来源'] = source[0].text if source else ''
        add_time = self.driver.xpath('//android.widget.TextView[starts-with(@text, "添加时间")]/../android.widget.TextView[2]').all()
        info_dict['添加时间'] = add_time[0].text if add_time else ''
        self.driver.press('back')   # 点击‘返回’按钮，返回至个人信息页面
        info_dict['对方是否是好友'] = self.is_friend()
        print(info_dict)
        if save_to_excel:
            self.save_to_excel(info_dict)
    
    def is_friend(self):
        """查询对方是否为你的好友,使用转账方式确认,如果是,会出现转账页面,否则会提示你不是ta的好友,然后点击知道了"""
        is_friend = '未知'
        self.driver(text='发消息').click()
        self.driver(description='更多功能按钮，已折叠').click()
        self.driver(text='转账').click()
        self.driver(resourceId='com.tencent.mm:id/keyboard_1').exists(timeout=3)
        self.driver(resourceId='com.tencent.mm:id/keyboard_1').click()
        self.driver(text='转账').click_gone()
        flag = self.driver(text='微信支付').wait_gone(timeout=5)
        if flag:
            if self.driver(textContains='付款方式').exists(timeout=1):
                self.driver.press('back')
                is_friend = '是'
            else:
                self.driver(textContains='知道了').click()
                is_friend = '否'

        while True:
            if self.driver(text='发消息').exists(timeout=0.2):
                break
            else:
                self.driver.press('back')
        return is_friend
    
    def save_to_excel(self, info_dict):
        """保存单条数据到Excel
        :param info_dict: 单条数据,是一个字典
        """
        
        titles = ['序号', '备注', '昵称', '微信号', '地区', '个性签名', '我和TA的共同群聊', '来源', '添加时间', '对方是否是好友', '日期', '头像']
        
        if os.path.isfile(self.excel_path):
            # 打开现有的工作簿
            wb = load_workbook(self.excel_path)
            ws = wb.active
            # 获取现有数据的行数
            last_row = ws.max_row + 1
        else:
            # 创建一个新的工作簿
            wb = Workbook()
            ws = wb.active
            # 写入表头
            ws.append(titles)
            last_row = 2

        data = [info_dict.get(key, '') for key in titles if key != '头像']

        # 写入数据
        ws.append(data)

        # 插入头像
        hp = info_dict.get('头像')
        if hp and os.path.isfile(hp):
            try:
                img = Image(hp)
                img.width = 100
                img.height = 100
                ws.add_image(img, f'L{last_row}')
            except Exception as e:
                print(f"插入图片时出错: {e}")

        # 自适应单元格大小
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # 改变头像列宽行高
        ws.column_dimensions['L'].width = 15
        ws.row_dimensions[last_row].height = 80

        # 设置所有单元格的文本居中显示
        for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=len(titles)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 保存文件
        wb.save(self.excel_path)

    def read_excel_names(self):
        """读取Excel中的备注，如果没有备注，就用昵称代替"""
        name_list = []
        workbook = load_workbook(self.excel_path)
        sheet = workbook.worksheets[0]
        column = sheet['B']
        for n, cell in enumerate(column):
            name = cell.value
            if name is None:
                name = sheet[f'C{n+1}'].value
            name_list.append(name)
        return name_list[1:]    # 去掉第一行表头"备注""
    

if __name__ == '__main__':
    @time_decorator
    def main():
        # TODO 新增从指定人员开始，也可以读取excel数据指定
        wechat = WeChat()
        all_person_name_list = []   # 所有人的名字列表
        wechat.start_wechat()
        myself_name = wechat.get_myself_name    # 获取自己的名字
        wechat.driver(text='通讯录').click()
        ignore_name_list = ['微信团队', '文件传输助手', myself_name]    # 忽略微信团队，文件传输助手，自己
        all_person = wechat.get_current_view_all_person()   # 获取当前页面的所有联系人
        # 获取第一个联系人
        current_person = all_person[0]
        index = 1

        if os.path.isfile(wechat.excel_path):   # 判断是否存在excel文件，如果存在，就从excel中读取备注
            name_list = wechat.read_excel_names()
            all_person_name_list = name_list
            index = len(all_person_name_list) + 1   # 获取当前列表的长度，作为下一次写入的序号
            # 滑动到最后name_list中最后一个人
            last_name = name_list[-1]
            while True:
                if wechat.driver(text=last_name).exists(timeout=0.1):
                    break
                else:
                    wechat.driver.swipe(0.5, 0.8, 0.5, 0.6, steps=20)
            time.sleep(0.5)
            all_person = wechat.get_current_view_all_person()
            if all_person[-1].text == last_name:
                wechat.driver.swipe(0.5, 0.8, 0.5, 0.7, steps=20)   # 如果当前屏幕上最后一个人等于excel中最后一个人，就滑动
                all_person = wechat.get_current_view_all_person()   # 重新获取

            for n, person in enumerate(all_person):
                if person.text == last_name:
                    try:
                        current_person = all_person[n+1]
                    except IndexError:
                        print('没有更多联系人了')
                    break  

        try:
            while True:
                try:
                    current_person_name = current_person.text
                except UiObjectNotFoundError:
                    break
                if current_person_name not in all_person_name_list:
                    if current_person_name not in ignore_name_list:
                        all_person_name_list.append(current_person_name)
                        try:
                            current_person.click()
                            wechat.get_person_info(index=index, is_friend=True, save_to_excel=True)
                            wechat.driver.press('back')
                        except Exception as e:
                            print(f'获取人员【{current_person_name}】信息失败: {e}')
                        finally:
                            index += 1

                    # 滑动当前联系人text的bounds['bottom']到屏幕0.7位置
                    wechat.driver.swipe(500, current_person.bounds[3], 500, wechat.driver.window_size()[1] * 0.7, steps=20)
                    # 刷新当前页面的所有联系人
                    all_person = wechat.get_current_view_all_person()
                    
                    for person in all_person:
                        person_name = person.text
                        if person_name not in all_person_name_list and person_name not in ignore_name_list:
                            current_person = person
                            break
                else:
                    break
        except KeyboardInterrupt:
            print('已停止当前的运行任务！')
        
    main()
